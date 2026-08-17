"""Deterministic parsers for retrieved supplementary-material files
(supplementary-material and structured-table retrieval layer).

Adapter-agnostic on purpose -- reusable by a future PANGAEA/BCO-DMO
supplement fetcher too, not Europe-PMC-specific. Mirrors
`sources/dna_extension.py`'s design: only a small, curated set of exact/
near-exact column-header aliases produce facts, never fuzzy matching, and
never one `RawFact` per raw cell keyed by an uncontrolled header string --
`mapping/rules.py`'s whole design rests on a small controlled vocabulary,
and per-cell facts on arbitrary spreadsheet headers would bloat `raw_facts`
with fact types nothing ever maps. Unrecognized columns are reported (name
and count) for a caller to log/inventory, never turned into facts.

Column-header aliases target this pipeline's own existing native names
that already have `mapping/rules.py` SAMPLE-level coverage (collection_date,
depth, lat_lon, temp, salinity, ph, samp_collect_device, samp_size, ...) -- a recognized column slots a supplement-derived fact into
the exact same downstream mapping path as any other structured source,
with no new mapping rules required for the common case.
"""
from __future__ import annotations

import csv
import io
import json
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass, field
from typing import Any
import re

from fair_ocean_agent.database.enums import EntityLevel, EntityRelationshipType, SupportType
from fair_ocean_agent.sources.base import EntityLinkCandidate, RawFactCandidate
from fair_ocean_agent.sources.replicate_grouping import ReplicateGroup, detect_replicate_groups

# Canonical native_name -> header variants that should map to it. Modest and
# curated rather than exhaustive -- expanding this table as real
# supplementary tables are observed is a natural, incremental follow-up
# (the same way sources/ncbi.py's BioSample attribute list grew from real
# observed data), not something to guess exhaustively up front.
SUPPLEMENT_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "collection_date": ("collection_date", "collection date", "date", "date_collected", "sampling_date", "sample_date"),
    "depth": ("depth", "depth_m", "depth (m)", "depth range (m)", "sampling_depth", "water_depth"),
    "elev": ("elev", "elevation", "altitude"),
    "geo_loc_name": ("geo_loc_name", "location", "locality", "site", "sites", "geographic_location"),
    "lat_lon": ("lat_lon", "lat/lon", "coordinates"),
    # "latitude"/"longitude" (plain, standard-agnostic) -- not FAIRe's own
    # "decimalLatitude"/"decimalLongitude" spelling. A raw fact's own
    # identity must never be a standard's field name (see extraction/
    # faire_fields.py's v2->v3 correction this session); mapping/rules.py
    # carries the correspondence to FAIRe's fields, not this alias table.
    "latitude": ("latitude", "lat", "lat.", "decimallatitude", "decimal_latitude"),
    "longitude": ("longitude", "lon", "long", "lon.", "decimallongitude", "decimal_longitude"),
    "samp_collect_device": ("samp_collect_device", "collection_device", "sampling_device", "sampler"),
    "samp_collect_method": ("samp_collect_method", "collection_method", "sampling_method"),
    "samp_size": ("samp_size", "sample_size", "sample_volume", "sample_mass"),
    "samp_size_unit": ("samp_size_unit", "sample_size_unit", "sample_volume_unit"),
    "temp": ("temp", "temperature", "water_temp", "water_temperature", "temp_c", "temperature (c)", "temperature (°c)"),
    "salinity": ("salinity", "sal", "salinity_psu", "salinity (psu)"),
    "ph": ("ph", "ph_value"),
    "chlorophyll": ("chlorophyll", "chl", "chl_a", "chlorophyll_a"),
    "env_broad_scale": ("env_broad_scale", "biome"),
    "env_local_scale": ("env_local_scale", "environmental_feature"),
    "env_medium": ("env_medium", "environmental_material"),
    "target_gene": ("target_gene", "gene", "marker", "markers name", "locus"),
    "target_subfragment": ("target_subfragment", "subfragment", "region", "hypervariable_region"),
    "assay_name": ("assay_name", "assay name", "assay", "assay_id", "assay id"),
    "pcr_plate_id": ("pcr_plate_id", "pcr plate id", "pcr_plate", "plate_id", "plate id"),
    "lib_id": ("lib_id", "library_id", "library id", "library", "lib name"),
    "seq_run_id": ("seq_run_id", "sequencing_run_id", "sequencing run id", "run_accession", "run accession"),
    "phix_perc": (
        "phix_perc", "phix_percentage", "phix percentage", "phix", "% phix", "phix%",
        "%phix", "phix_pct", "percent phix",
    ),
    "filename": ("filename", "file_name", "fastq_1", "fastq file 1", "forward_reads", "r1"),
    "filename2": ("filename2", "file_name2", "fastq_2", "fastq file 2", "reverse_reads", "r2"),
    "checksum_filename": ("checksum_filename", "checksum_file_1", "md5_1", "fastq_md5_1"),
    "checksum_filename2": ("checksum_filename2", "checksum_file_2", "md5_2", "fastq_md5_2"),
    "associatedSequences": ("associatedSequences", "associated_sequences", "associated sequences"),
    "input_read_count": ("input_read_count", "read_count", "read count", "reads", "raw_reads", "raw read count"),
    "platform": ("platform", "instrument_platform", "sequencing_platform", "sequencing platform"),
}

def _normalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", header.lower()).strip()


_ALIAS_TO_CANONICAL = {
    _normalize_header(alias): canonical for canonical, aliases in SUPPLEMENT_COLUMN_ALIASES.items() for alias in aliases
}

# Header variants signaling "this column identifies the sample/run this row
# is about" -- when present, each row's facts bind to that entity (get-or-
# create happens later, at persistence time, exactly like every other
# adapter's RawFactCandidate.entity_external_id already works) rather than
# broadcasting study-wide.
_SAMPLE_IDENTIFIER_ALIASES = frozenset(
    {
        _normalize_header(alias)
        for alias in (
            "sample_id",
            "sample_name",
            "sample",
            "biosample_accession",
            "specimen_id",
            "specimen",
            "station",
            "station_id",
        )
    }
)
_RUN_IDENTIFIER_ALIASES = frozenset(
    _normalize_header(alias) for alias in ("run_accession", "run accession", "run_id", "sra_run", "seq_run_id")
)

_EXPERIMENT_RUN_FACTS = frozenset(
    {
        "assay_name",
        "pcr_plate_id",
        "lib_id",
        "seq_run_id",
        "phix_perc",
        "filename",
        "filename2",
        "checksum_filename",
        "checksum_filename2",
        "associatedSequences",
        "input_read_count",
        "platform",
    }
)


@dataclass
class ParsedTableResult:
    facts: list[RawFactCandidate] = field(default_factory=list)
    recognized_columns: list[str] = field(default_factory=list)
    unrecognized_columns: list[str] = field(default_factory=list)
    row_count: int = 0

    def summary(self) -> str:
        parts = [f"{self.row_count} row(s)"]
        if self.recognized_columns:
            parts.append(f"recognized: {', '.join(self.recognized_columns)}")
        if self.unrecognized_columns:
            parts.append(f"unrecognized: {', '.join(self.unrecognized_columns)}")
        return "; ".join(parts)


def _column_letter(index: int) -> str:
    """0-indexed column number -> spreadsheet-style letter (0->A, 25->Z, 26->AA)."""
    letters = ""
    index += 1
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def _entity_binding(header_lower: str) -> EntityLevel | None:
    if header_lower in _SAMPLE_IDENTIFIER_ALIASES:
        return EntityLevel.SAMPLE
    if header_lower in _RUN_IDENTIFIER_ALIASES:
        return EntityLevel.SEQUENCING_RUN
    return None


def _facts_from_rows(
    rows: list[list[str]],
    file_name: str,
    sheet_name: str | None = None,
) -> ParsedTableResult:
    """Shared row-processing for any tabular source (CSV/TSV/XLSX/XLS) once
    reduced to a uniform list-of-rows-of-strings, header row first."""
    result = ParsedTableResult()
    if not rows:
        return result

    # Real supplementary spreadsheets commonly carry a caption/title row
    # ("Supplementary Table S5. Accession Number.") above the actual header
    # row (confirmed live against a real PMC article's Table_5.xlsx) --
    # a caption row has at most one non-blank cell, unlike a real header row
    # which names >=2 columns. Skip any such leading rows so the real header
    # (and its data) don't get silently misread as an unrecognized caption.
    header_index = 0
    while header_index < len(rows) - 1 and sum(1 for cell in rows[header_index] if str(cell).strip()) < 2:
        header_index += 1

    header = [str(cell).strip() if cell is not None else "" for cell in rows[header_index]]
    data_rows = rows[header_index + 1 :]
    result.row_count = len(data_rows)

    locator_prefix = f"supplement.{file_name}"
    if sheet_name:
        locator_prefix = f"{locator_prefix}#{sheet_name}"

    # Every real header cell, verbatim, pipe-joined, regardless of whether
    # it's recognized below -- a cheap diagnostic (projectMetadata.
    # spreadsheet_headers) so a human can see what a supplement table
    # actually contains even when none of its columns are in
    # SUPPLEMENT_COLUMN_ALIASES yet, per an explicit user request. Built
    # here but appended to result.facts at every return point below (kept
    # LAST, not first) so it never disturbs an existing caller's own
    # facts[0]-is-the-first-real-fact assumption; unconditional (fires on
    # the "nothing else recognized" early-return path too), since an
    # all-unrecognized header row is exactly the case this exists to
    # surface.
    non_empty_headers = [h for h in header if h]
    spreadsheet_headers_fact = (
        RawFactCandidate(
            entity_level=EntityLevel.STUDY,
            fact_type_candidate="spreadsheet_headers",
            raw_field_name="header_row",
            raw_value=" | ".join(non_empty_headers),
            source_locator=f"{locator_prefix}!row{header_index + 1}",
        )
        if non_empty_headers
        else None
    )

    column_plan: list[tuple[int, str, str | None]] = []  # (col_index, canonical, header_text)
    identifier_cols: dict[EntityLevel, tuple[int, str]] = {}
    for col_index, raw_header in enumerate(header):
        if not raw_header:
            continue
        lowered = _normalize_header(raw_header)
        entity_level = _entity_binding(lowered)
        if entity_level is not None and entity_level not in identifier_cols:
            identifier_cols[entity_level] = (col_index, raw_header)
            continue
        canonical = _ALIAS_TO_CANONICAL.get(lowered)
        if canonical:
            column_plan.append((col_index, canonical, raw_header))
            result.recognized_columns.append(raw_header)
        else:
            result.unrecognized_columns.append(raw_header)

    replicate_group_by_sample_id: dict[str, ReplicateGroup] = {}
    if EntityLevel.SAMPLE in identifier_cols:
        sample_id_index, _ = identifier_cols[EntityLevel.SAMPLE]
        sample_names_by_id = {
            str(row[sample_id_index]).strip(): str(row[sample_id_index]).strip()
            for row in data_rows
            if sample_id_index < len(row) and row[sample_id_index] not in (None, "")
        }
        replicate_group_by_sample_id = {
            member: group
            for group in detect_replicate_groups(sample_names_by_id)
            for member in group.members
        }

    if not column_plan and not replicate_group_by_sample_id:
        if spreadsheet_headers_fact is not None:
            result.facts.append(spreadsheet_headers_fact)
        return result

    for row_offset, row in enumerate(data_rows):
        row_number = header_index + row_offset + 2  # +1 for 1-indexing, +1 for the header row
        row_identifiers: dict[EntityLevel, str] = {}
        for id_level, (id_index, _) in identifier_cols.items():
            if id_index < len(row) and row[id_index] not in (None, ""):
                row_identifiers[id_level] = str(row[id_index]).strip()

        canonical_values = {
            canonical: str(row[col_index]).strip()
            for col_index, canonical, _raw_header in column_plan
            if col_index < len(row) and row[col_index] not in (None, "")
        }
        has_experiment_metadata = any(name in _EXPERIMENT_RUN_FACTS for name in canonical_values)
        experiment_id = canonical_values.get("lib_id")
        if has_experiment_metadata and not experiment_id:
            sheet_part = f":{sheet_name}" if sheet_name else ""
            experiment_id = f"internal:supplement:{file_name}{sheet_part}:row:{row_number}"

        experiment_links: list[EntityLinkCandidate] = []
        sample_id = row_identifiers.get(EntityLevel.SAMPLE)
        run_id = row_identifiers.get(EntityLevel.SEQUENCING_RUN) or canonical_values.get("seq_run_id")
        assay_name = canonical_values.get("assay_name")
        if sample_id:
            experiment_links.append(
                EntityLinkCandidate(
                    entity_level=EntityLevel.SAMPLE,
                    external_identifier=sample_id,
                    relationship_type=EntityRelationshipType.DERIVED_FROM_SAMPLE,
                    label=sample_id,
                )
            )
        if run_id:
            experiment_links.append(
                EntityLinkCandidate(
                    entity_level=EntityLevel.SEQUENCING_RUN,
                    external_identifier=run_id,
                    relationship_type=EntityRelationshipType.SEQUENCED_IN_RUN,
                    label=run_id,
                )
            )
        if assay_name:
            experiment_links.append(
                EntityLinkCandidate(
                    entity_level=EntityLevel.ASSAY,
                    external_identifier=assay_name,
                    relationship_type=EntityRelationshipType.USES_ASSAY,
                    label=assay_name,
                )
            )

        for col_index, canonical, raw_header in column_plan:
            if col_index >= len(row):
                continue
            value = row[col_index]
            if value in (None, ""):
                continue
            entity_level = EntityLevel.STUDY
            entity_external_id = None
            entity_links: list[EntityLinkCandidate] = []
            if canonical in _EXPERIMENT_RUN_FACTS and experiment_id:
                entity_level = EntityLevel.EXPERIMENT_RUN
                entity_external_id = experiment_id
                entity_links = experiment_links
            elif EntityLevel.SAMPLE in row_identifiers:
                entity_level = EntityLevel.SAMPLE
                entity_external_id = row_identifiers[EntityLevel.SAMPLE]
            elif EntityLevel.SEQUENCING_RUN in row_identifiers:
                entity_level = EntityLevel.SEQUENCING_RUN
                entity_external_id = row_identifiers[EntityLevel.SEQUENCING_RUN]
            cell_ref = f"{_column_letter(col_index)}{row_number}"
            result.facts.append(
                RawFactCandidate(
                    entity_level=entity_level,
                    fact_type_candidate=canonical,
                    raw_field_name=raw_header,
                    raw_value=str(value).strip(),
                    source_locator=f"{locator_prefix}!{cell_ref}",
                    entity_external_id=entity_external_id,
                    entity_label=entity_external_id,
                    entity_links=entity_links,
                )
            )
        if experiment_id and sample_id:
            id_index, raw_header = identifier_cols[EntityLevel.SAMPLE]
            result.facts.append(
                RawFactCandidate(
                    entity_level=EntityLevel.EXPERIMENT_RUN,
                    fact_type_candidate="samp_name",
                    raw_field_name=raw_header,
                    raw_value=sample_id,
                    source_locator=f"{locator_prefix}!{_column_letter(id_index)}{row_number}",
                    entity_external_id=experiment_id,
                    entity_label=experiment_id,
                    entity_links=experiment_links,
                )
            )
        if experiment_id and run_id and "seq_run_id" not in canonical_values:
            id_index, raw_header = identifier_cols[EntityLevel.SEQUENCING_RUN]
            result.facts.append(
                RawFactCandidate(
                    entity_level=EntityLevel.EXPERIMENT_RUN,
                    fact_type_candidate="seq_run_id",
                    raw_field_name=raw_header,
                    raw_value=run_id,
                    source_locator=f"{locator_prefix}!{_column_letter(id_index)}{row_number}",
                    entity_external_id=experiment_id,
                    entity_label=experiment_id,
                    entity_links=experiment_links,
                )
            )
        if sample_id:
            group = replicate_group_by_sample_id.get(sample_id)
            if group is not None:
                id_index, raw_header = identifier_cols[EntityLevel.SAMPLE]
                result.facts.append(
                    RawFactCandidate(
                        entity_level=EntityLevel.SAMPLE,
                        fact_type_candidate="biological_rep_relation",
                        raw_field_name=raw_header,
                        raw_value=" | ".join(group.members),
                        source_locator=f"{locator_prefix}!{_column_letter(id_index)}{row_number}",
                        entity_external_id=sample_id,
                        entity_label=sample_id,
                        support_type=SupportType.DETERMINISTICALLY_DERIVED,
                        confidence_metadata={
                            "replicate_detection_signal": group.signal.value,
                            "replicate_group_size": len(group.members),
                        },
                    )
                )
    if spreadsheet_headers_fact is not None:
        result.facts.append(spreadsheet_headers_fact)
    return result


def parse_delimited_table(content: bytes, file_name: str, delimiter: str = ",") -> ParsedTableResult:
    """CSV/TSV. Decodes as UTF-8 with a latin-1 fallback (real supplementary
    files from non-English-first journals aren't reliably UTF-8)."""
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader if any(cell.strip() for cell in row)]
    return _facts_from_rows(rows, file_name)


def parse_xlsx_table(content: bytes, file_name: str) -> list[ParsedTableResult]:
    """One ParsedTableResult per sheet -- the user's own example
    ("Supplementary_Table_2.xlsx, sheet sample_metadata, cell H42") expects
    sheet-level locators, so every sheet is parsed and reported separately
    rather than merged."""
    import openpyxl

    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    results = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = [[cell if cell is not None else "" for cell in row] for row in sheet.iter_rows(values_only=True)]
        rows = [row for row in rows if any(str(cell).strip() for cell in row)]
        results.append(_facts_from_rows(rows, file_name, sheet_name=sheet_name))
    workbook.close()
    return results


def parse_xls_table(content: bytes, file_name: str) -> list[ParsedTableResult]:
    """Legacy .xls (openpyxl only reads .xlsx) -- same per-sheet contract as
    parse_xlsx_table."""
    import xlrd

    workbook = xlrd.open_workbook(file_contents=content)
    results = []
    for sheet in workbook.sheets():
        rows = [
            [sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)]
            for row_index in range(sheet.nrows)
        ]
        rows = [row for row in rows if any(str(cell).strip() for cell in row)]
        results.append(_facts_from_rows(rows, file_name, sheet_name=sheet.name))
    return results


def _iter_structured_fields(value: Any, path: str) -> Iterable[tuple[str, str, Any]]:
    """Same recursive walk as dna_extension.py's _iter_fields, generalized
    for any JSON-like nested structure (dict/list of scalars)."""
    if isinstance(value, dict):
        for key, child in value.items():
            child_path = f"{path}.{key}" if path else str(key)
            yield key, child_path, child
            yield from _iter_structured_fields(child, child_path)
    elif isinstance(value, list):
        for index, child in enumerate(value):
            child_path = f"{path}[{index}]"
            yield from _iter_structured_fields(child, child_path)


def parse_json_supplement(content: bytes, file_name: str) -> ParsedTableResult:
    result = ParsedTableResult()
    try:
        payload = json.loads(content)
    except (json.JSONDecodeError, UnicodeDecodeError):
        return result

    recognized: set[str] = set()
    unrecognized: set[str] = set()
    for raw_key, path, value in _iter_structured_fields(payload, f"supplement.{file_name}"):
        if value in (None, "", [], {}) or isinstance(value, (dict, list)):
            continue
        canonical = _ALIAS_TO_CANONICAL.get(str(raw_key).strip().lower())
        if canonical is None:
            unrecognized.add(str(raw_key))
            continue
        recognized.add(str(raw_key))
        result.facts.append(
            RawFactCandidate(
                entity_level=EntityLevel.STUDY,
                fact_type_candidate=canonical,
                raw_field_name=str(raw_key),
                raw_value=str(value).strip(),
                source_locator=path,
            )
        )
    result.recognized_columns = sorted(recognized)
    result.unrecognized_columns = sorted(unrecognized)
    return result


def parse_xml_supplement(content: bytes, file_name: str) -> ParsedTableResult:
    result = ParsedTableResult()
    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        return result

    recognized: set[str] = set()
    unrecognized: set[str] = set()
    for index, element in enumerate(root.iter()):
        tag = element.tag.rsplit("}", 1)[-1]  # strip any XML namespace
        text = (element.text or "").strip()
        if not text or list(element):
            continue  # only leaf elements with direct text content are candidate facts
        canonical = _ALIAS_TO_CANONICAL.get(tag.lower())
        if canonical is None:
            unrecognized.add(tag)
            continue
        recognized.add(tag)
        result.facts.append(
            RawFactCandidate(
                entity_level=EntityLevel.STUDY,
                fact_type_candidate=canonical,
                raw_field_name=tag,
                raw_value=text,
                source_locator=f"supplement.{file_name}#{index}.{tag}",
            )
        )
    result.recognized_columns = sorted(recognized)
    result.unrecognized_columns = sorted(unrecognized)
    return result


def parse_zip_supplement(content: bytes, file_name: str, max_member_bytes: int) -> list[ParsedTableResult]:
    """Parse supported structured files inside a supplementary ZIP.

    Nested archives, PDFs, office documents, and images are intentionally
    skipped here: this function is the deterministic structured-table path,
    not the LLM/text path. Each inner file name is included in source
    locators as ``outer.zip!inner.xlsx`` so provenance remains precise.
    """
    results: list[ParsedTableResult] = []
    with zipfile.ZipFile(io.BytesIO(content)) as zip_file:
        for info in zip_file.infolist():
            if info.is_dir():
                continue
            inner_name = info.filename
            extension = inner_name.rsplit(".", 1)[-1].lower() if "." in inner_name else ""
            if extension not in {"csv", "tsv", "xlsx", "xls", "json", "xml"}:
                continue
            inner_content = safe_read_zip_member(zip_file, info, max_member_bytes)
            locator_name = f"{file_name}!{inner_name}"
            if extension == "csv":
                results.append(parse_delimited_table(inner_content, locator_name, delimiter=","))
            elif extension == "tsv":
                results.append(parse_delimited_table(inner_content, locator_name, delimiter="\t"))
            elif extension == "xlsx":
                results.extend(parse_xlsx_table(inner_content, locator_name))
            elif extension == "xls":
                results.extend(parse_xls_table(inner_content, locator_name))
            elif extension == "json":
                results.append(parse_json_supplement(inner_content, locator_name))
            elif extension == "xml":
                results.append(parse_xml_supplement(inner_content, locator_name))
    return results


def extract_pdf_text(content: bytes) -> str:
    """Best-effort text extraction for a PDF supplementary-methods
    document -- handed to the *existing* LLM extraction path
    (extraction/text.py:extract_facts_from_section) by the caller, never
    parsed for facts here directly (a PDF's prose has no controlled
    vocabulary to alias-match against)."""
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    return "\n\n".join(page.extract_text() or "" for page in reader.pages).strip()


_DOCX_WORD_NAMESPACE = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
_DOCX_PARAGRAPH_TAG = f"{_DOCX_WORD_NAMESPACE}p"
_DOCX_TEXT_TAG = f"{_DOCX_WORD_NAMESPACE}t"


def extract_docx_text(content: bytes) -> str:
    """Best-effort text extraction for a .docx supplementary document --
    a real gap found live: Frontiers routinely ships its "Data Sheet"
    supplements as .docx, and this extension previously had no handler at
    all, falling into the generic "unsupported file type, not parsed"
    branch, so a real paper's supplement was silently never scanned for
    anything (controls included) despite being successfully retrieved.

    A .docx file is a zip archive; `word/document.xml` holds the visible
    text as `<w:t>` runs inside `<w:p>` paragraphs. Deliberately dependency-
    free (stdlib `zipfile`/`xml.etree.ElementTree` only, matching this
    module's existing JATS/XML parsing) rather than adding python-docx as
    a new dependency for what is, at its core, just zipped XML."""
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        try:
            document_xml = archive.read("word/document.xml")
        except KeyError:
            return ""
    root = ET.fromstring(document_xml)
    paragraphs = []
    for paragraph in root.iter(_DOCX_PARAGRAPH_TAG):
        text = "".join(node.text or "" for node in paragraph.iter(_DOCX_TEXT_TAG))
        if text.strip():
            paragraphs.append(text)
    return "\n\n".join(paragraphs).strip()


class ZipMemberTooLargeError(Exception):
    """Raised when a zip member's uncompressed size exceeds the configured
    per-member cap -- either as reported by ZipInfo.file_size (checked
    before ever reading) or discovered mid-stream (a zip-bomb defense: never
    trust the reported size alone)."""


def safe_read_zip_member(zip_file: zipfile.ZipFile, info: zipfile.ZipInfo, max_bytes: int) -> bytes:
    """Reads one zip member with a hard uncompressed-size cap, checked both
    from the reported metadata (cheap, catches the common case) and while
    streaming (defense in depth against a mismatched/malicious reported
    size -- a real zip-bomb concern for an all-supplementary-files bundle
    this pipeline doesn't control the contents of). Note zipfile's own CRC
    validation independently catches a corrupted/understated `file_size`
    too (raises `BadZipFile` before this function's own streaming check
    would); the streaming check here specifically guards the case where
    the *reported* size is accurate but simply exceeds the cap, checked
    incrementally rather than only once upfront."""
    if info.file_size > max_bytes:
        raise ZipMemberTooLargeError(f"{info.filename}: reported size {info.file_size} exceeds cap {max_bytes}")

    chunks: list[bytes] = []
    total = 0
    with zip_file.open(info) as member:
        while True:
            chunk = member.read(65536)
            if not chunk:
                break
            total += len(chunk)
            if total > max_bytes:
                raise ZipMemberTooLargeError(
                    f"{info.filename}: exceeded cap {max_bytes} while streaming (reported size was misleading)"
                )
            chunks.append(chunk)
    return b"".join(chunks)
