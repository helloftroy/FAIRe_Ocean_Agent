"""Tests for sources/supplement_parsing.py: deterministic CSV/XLSX/XLS/
JSON/XML parsing (alias-matched columns only, per-cell provenance), the
zip-bomb-safe extraction helper, and PDF text extraction."""
import io
import json
import zipfile

import openpyxl
import pytest

from fair_ocean_agent.database.enums import EntityLevel
from fair_ocean_agent.sources.supplement_parsing import (
    ZipMemberTooLargeError,
    _column_letter,
    extract_pdf_text,
    parse_delimited_table,
    parse_json_supplement,
    parse_xls_table,
    parse_xlsx_table,
    parse_xml_supplement,
    safe_read_zip_member,
)


def test_column_letter_handles_multi_letter_columns():
    assert _column_letter(0) == "A"
    assert _column_letter(25) == "Z"
    assert _column_letter(26) == "AA"
    assert _column_letter(27) == "AB"


def test_parse_delimited_table_only_recognizes_aliased_columns():
    csv_content = b"sample_id,collection_date,temp,notes\nS1,2023-01-05,18.2,fine\nS2,2023-01-06,17.9,cloudy\n"
    result = parse_delimited_table(csv_content, "Supplementary_Table_1.csv")

    assert result.row_count == 2
    assert sorted(result.recognized_columns) == ["collection_date", "temp"]
    assert result.unrecognized_columns == ["notes"]
    fact_types = {f.fact_type_candidate for f in result.facts}
    assert fact_types == {"collection_date", "temp"}
    assert "notes" not in fact_types


def test_parse_delimited_table_binds_recognized_identifier_column_to_sample_entity():
    csv_content = b"sample_id,temp\nS1,18.2\n"
    result = parse_delimited_table(csv_content, "t.csv")
    fact = result.facts[0]
    assert fact.entity_level == EntityLevel.SAMPLE
    assert fact.entity_external_id == "S1"
    assert fact.entity_label == "S1"


def test_parse_delimited_table_defaults_to_study_level_without_identifier_column():
    csv_content = b"temp,salinity\n18.2,35\n"
    result = parse_delimited_table(csv_content, "t.csv")
    assert all(f.entity_level == EntityLevel.STUDY for f in result.facts)
    assert all(f.entity_external_id is None for f in result.facts)


def test_parse_delimited_table_source_locator_matches_the_users_own_example_format():
    csv_content = b"sample_id,temp\nS1,18.2\n"
    result = parse_delimited_table(csv_content, "Supplementary_Table_2.csv")
    assert result.facts[0].source_locator == "supplement.Supplementary_Table_2.csv!B2"


def test_parse_delimited_table_skips_a_leading_caption_row():
    """Regression guard for a real bug found via live validation against
    PMC7469538's Table_5.xlsx: a caption/title row above the real header
    row (only one non-blank cell) must not be misread as the header --
    otherwise the real header+data rows are silently misaligned and every
    column comes back unrecognized."""
    csv_content = b"Supplementary Table S5. Accession Number.,,\nsample_id,temp\nS1,18.2\n"
    result = parse_delimited_table(csv_content, "t.csv")
    assert result.row_count == 1
    assert {f.fact_type_candidate for f in result.facts} == {"temp"}
    assert result.facts[0].entity_external_id == "S1"
    assert result.facts[0].source_locator == "supplement.t.csv!B3"


def test_parse_delimited_table_recognizes_real_supplement_environment_headers():
    csv_content = (
        b"station,Sites,Lat.,Lon.,Depth range (m),Markers name\n"
        b"MC751,Seep site,27.1,-91.2,430-450,16S rRNA\n"
    )

    result = parse_delimited_table(csv_content, "gcb_sites.csv")
    fact_types = {f.fact_type_candidate for f in result.facts}

    assert fact_types == {"geo_loc_name", "latitude", "longitude", "depth", "target_gene"}
    assert all(f.entity_external_id == "MC751" for f in result.facts)


def test_parse_delimited_table_skips_blank_cells():
    csv_content = b"sample_id,temp,ph\nS1,,8.1\n"
    result = parse_delimited_table(csv_content, "t.csv")
    fact_types = {f.fact_type_candidate for f in result.facts}
    assert fact_types == {"ph"}


def test_parse_delimited_table_handles_tsv_delimiter():
    tsv_content = b"sample_id\ttemp\nS1\t18.2\n"
    result = parse_delimited_table(tsv_content, "t.tsv", delimiter="\t")
    assert result.row_count == 1
    assert result.facts[0].raw_value == "18.2"


def test_parse_delimited_table_falls_back_to_latin1_on_bad_utf8():
    content = "sample_id,notes\nS1,café\n".encode("latin-1")
    result = parse_delimited_table(content, "t.csv")
    assert result.row_count == 1


def _xlsx_bytes(rows: list[list], sheet_name: str = "Sheet1") -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def test_parse_xlsx_table_uses_native_not_faire_names_for_coordinates():
    """Regression guard: latitude/longitude must be this pipeline's own
    standard-agnostic native names, never FAIRe's own decimalLatitude/
    decimalLongitude spelling (a raw fact's identity must never be a
    standard's field name)."""
    content = _xlsx_bytes([["sample_id", "Latitude", "Longitude"], ["S1", 34.05, -118.25]], sheet_name="sample_metadata")
    results = parse_xlsx_table(content, "Supplementary_Table_2.xlsx")
    assert len(results) == 1
    fact_types = {f.fact_type_candidate for f in results[0].facts}
    assert fact_types == {"latitude", "longitude"}
    assert "decimalLatitude" not in fact_types


def test_parse_xlsx_table_locator_includes_sheet_name_and_cell():
    content = _xlsx_bytes([["sample_id", "ph"], ["S1", 8.1]], sheet_name="sample_metadata")
    results = parse_xlsx_table(content, "Supplementary_Table_2.xlsx")
    fact = results[0].facts[0]
    assert fact.source_locator == "supplement.Supplementary_Table_2.xlsx#sample_metadata!B2"


def test_parse_xlsx_table_one_result_per_sheet():
    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title = "sheet_a"
    sheet1.append(["sample_id", "temp"])
    sheet1.append(["S1", 18.2])
    sheet2 = workbook.create_sheet("sheet_b")
    sheet2.append(["sample_id", "salinity"])
    sheet2.append(["S2", 35])
    buf = io.BytesIO()
    workbook.save(buf)

    results = parse_xlsx_table(buf.getvalue(), "t.xlsx")
    assert len(results) == 2
    assert {f.fact_type_candidate for f in results[0].facts} == {"temp"}
    assert {f.fact_type_candidate for f in results[1].facts} == {"salinity"}


def test_parse_xls_table_via_fake_xlrd_workbook(monkeypatch):
    """xlrd's own binary format is awkward to synthesize in a test without
    an extra write-side dependency (xlwt) -- this exercises our own
    row-extraction logic against a minimal fake workbook exposing the same
    read interface (.sheets(), sheet.nrows/.ncols/.cell_value, .name) xlrd
    itself provides, rather than testing xlrd's own correctness."""
    class FakeSheet:
        name = "sample_metadata"
        nrows = 2
        ncols = 2
        _data = [["sample_id", "temp"], ["S1", 18.2]]

        def cell_value(self, row, col):
            return self._data[row][col]

    class FakeWorkbook:
        def sheets(self):
            return [FakeSheet()]

    import sys
    import types

    fake_xlrd_module = types.ModuleType("xlrd")
    fake_xlrd_module.open_workbook = lambda file_contents: FakeWorkbook()
    monkeypatch.setitem(sys.modules, "xlrd", fake_xlrd_module)

    results = parse_xls_table(b"irrelevant", "Table.xls")
    assert len(results) == 1
    assert results[0].facts[0].fact_type_candidate == "temp"
    assert results[0].facts[0].entity_external_id == "S1"


def test_parse_json_supplement_only_recognizes_aliased_keys():
    payload = json.dumps({"samples": [{"sample_id": "S1", "temp": "18.2", "notes": "x"}]}).encode()
    result = parse_json_supplement(payload, "meta.json")
    fact_types = {f.fact_type_candidate for f in result.facts}
    assert fact_types == {"temp"}
    assert "notes" in result.unrecognized_columns


def test_parse_json_supplement_returns_empty_on_invalid_json():
    result = parse_json_supplement(b"not json", "meta.json")
    assert result.facts == []


def test_parse_xml_supplement_only_recognizes_aliased_leaf_tags():
    xml = b"<root><record><temp>18.2</temp><notes>hi</notes></record></root>"
    result = parse_xml_supplement(xml, "meta.xml")
    fact_types = {f.fact_type_candidate for f in result.facts}
    assert fact_types == {"temp"}
    assert "notes" in result.unrecognized_columns


def test_parse_xml_supplement_returns_empty_on_malformed_xml():
    result = parse_xml_supplement(b"<not><valid", "meta.xml")
    assert result.facts == []


def test_extract_pdf_text_runs_without_error_on_a_real_pdf():
    from pypdf import PdfWriter

    writer = PdfWriter()
    writer.add_blank_page(width=200, height=200)
    buf = io.BytesIO()
    writer.write(buf)
    text = extract_pdf_text(buf.getvalue())
    assert isinstance(text, str)


def _zip_with(name: str, content: bytes) -> zipfile.ZipFile:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr(name, content)
    buf.seek(0)
    return zipfile.ZipFile(buf)


def test_safe_read_zip_member_returns_content_within_cap():
    zf = _zip_with("small.txt", b"hello world")
    info = zf.getinfo("small.txt")
    assert safe_read_zip_member(zf, info, max_bytes=1000) == b"hello world"
    zf.close()


def test_safe_read_zip_member_raises_on_reported_size_over_cap():
    zf = _zip_with("big.txt", b"x" * 1000)
    info = zf.getinfo("big.txt")
    with pytest.raises(ZipMemberTooLargeError):
        safe_read_zip_member(zf, info, max_bytes=500)
    zf.close()

