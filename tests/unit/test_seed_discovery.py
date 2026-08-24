from __future__ import annotations

from pathlib import Path
import gzip
import re
from zipfile import ZIP_DEFLATED, ZipFile

import httpx
from openpyxl import Workbook

from fair_ocean_agent.seed_discovery.clients.ena import EnaXrefClient
from fair_ocean_agent.seed_discovery.clients.ena_portal import EnaPortalClient, classify_sequence_accessibility, parse_read_run
from fair_ocean_agent.seed_discovery.clients.europepmc import EuropePmcSeedClient
from fair_ocean_agent.seed_discovery.clients.http import CachedHttpClient
from fair_ocean_agent.seed_discovery.clients.mgnify import MgnifyClient, extract_insdc_identifiers, parse_study
from fair_ocean_agent.seed_discovery.clients.openalex import OpenAlexSeedClient
from fair_ocean_agent.seed_discovery.config import SeedDiscoveryConfig
from fair_ocean_agent.seed_discovery.db import SeedDiscoveryDB, choose_primary_candidate, normalize_doi_for_seed
from fair_ocean_agent.seed_discovery.ena_discovery import aggregate_ena_study, build_ena_query_partitions
from fair_ocean_agent.seed_discovery.filters import is_marine_study
from fair_ocean_agent.seed_discovery.jgi_gold import inspect_gold_snapshot, process_gold_snapshot
from fair_ocean_agent.seed_discovery.local_epmc import DatasetAccessions, LocalEuropePmcResolver, normalize_epmc_accession
from fair_ocean_agent.seed_discovery.models import EnaRun, MatchConfidence, MgnifyStudy, PublicationCandidate, ResolutionStatus
from fair_ocean_agent.seed_discovery.publication_resolver import OpenAlexRateLimitError, PublicationResolver, title_similarity
from fair_ocean_agent.seed_discovery.update_epmc_accession_index import parse_id_mapping


def _db(tmp_path: Path) -> SeedDiscoveryDB:
    db = SeedDiscoveryDB(tmp_path / "seeds.sqlite")
    db.initialize()
    return db


def test_seed_discovery_doi_normalization():
    assert normalize_doi_for_seed("https://doi.org/10.1234/ABC.") == "10.1234/abc"
    assert normalize_doi_for_seed("DOI: 10.5555/Foo") == "10.5555/foo"


def test_seed_discovery_openalex_default_interval_is_conservative():
    assert SeedDiscoveryConfig().request_interval_for_source("openalex") == 10.0


def test_seed_discovery_openalex_disabled_by_default():
    assert SeedDiscoveryConfig().openalex_enabled is False


def test_mgnify_parser_extracts_insdc_identifiers_from_payload_text():
    payload = {
        "accession": "MGYS00000001",
        "study_name": "Marine sediment microbiome PRJEB12345",
        "description": "Submitted as ERP999999.",
        "biome": "root:Environmental:Aquatic:Marine",
    }

    assert extract_insdc_identifiers(payload) == ("PRJEB12345", "ERP999999")
    study = parse_study(payload)
    assert study.mgnify_accession == "MGYS00000001"
    assert study.bioproject_accession == "PRJEB12345"


def test_mgnify_parser_reads_real_nested_biome_object_shape():
    """Confirmed live against a real cached MGnify v2 /studies/ response:
    biome is a nested object ({"biome_name": "Soil", "lineage":
    "root:Environmental:Terrestrial:Soil"}), not a flat string.
    _first_string only ever matches a plain string, so before this test
    (and its matching fix in parse_study), biome silently stayed empty
    for every single study, which in turn meant is_marine_study's
    rejected_biome_terms check (see filters.py) never saw the "human"/
    "animal"/etc signal it needs -- confirmed live, several real
    Host-associated:Human studies (salivary/lung microbiome) incorrectly
    passed the filter as a direct result, entirely independent of the
    filter's own intentionally-broad accept policy (see
    test_broad_biome_filter_is_reject_only_by_default below)."""
    payload = {
        "accession": "MGYS00002209",
        "biome": {"biome_name": "Saliva", "lineage": "root:Host-associated:Human:Digestive system:Oral:Saliva"},
        "title": "Identification of salivary microbiota associated with oral malodor using 16S pyrosequencing",
    }
    study = parse_study(payload)
    assert study.biome == "Saliva (root:Host-associated:Human:Digestive system:Oral:Saliva)"
    assert not is_marine_study(study, SeedDiscoveryConfig())


def test_broad_biome_filter_is_reject_only_by_default():
    config = SeedDiscoveryConfig()

    assert is_marine_study(
        MgnifyStudy(mgnify_accession="MGYS1", biome="root:Environmental:Aquatic:Marine sediment"),
        config,
    )
    assert is_marine_study(
        MgnifyStudy(mgnify_accession="MGYS2", biome="root:Environmental:Aquatic:Freshwater sediment"),
        config,
    )
    assert is_marine_study(
        MgnifyStudy(mgnify_accession="MGYS3", biome="root:Environmental:Terrestrial:Soil"),
        config,
    )
    assert not is_marine_study(
        MgnifyStudy(mgnify_accession="MGYS4", biome="root:Host-associated:Human gut"),
        config,
    )
    assert not is_marine_study(
        MgnifyStudy(mgnify_accession="MGYS5", biome="root:Host-associated:Animal rumen"),
        config,
    )


def test_ena_query_partitions_include_primary_tags_and_optional_secondary_routes():
    primary = build_ena_query_partitions(SeedDiscoveryConfig(), include_secondary=False)
    secondary = build_ena_query_partitions(SeedDiscoveryConfig(), include_secondary=True)

    assert any('environment_biome="marine"' in part.query for part in primary)
    assert any('environment_biome="coral reef"' in part.query for part in primary)
    assert not any(part.query.startswith("tax_tree(") for part in primary)
    assert any(part.query == "tax_tree(408172)" for part in secondary)
    assert any("environment_material" in part.query and "coral reef" in part.query for part in secondary)


def test_ena_query_partitions_can_be_date_sharded():
    config = SeedDiscoveryConfig(ena_date_shards_enabled=True, ena_date_shard_start_year=2026)
    partitions = build_ena_query_partitions(config, include_secondary=False)

    assert any("first_public_" in part.name for part in partitions)
    assert any("first_public>=" in part.query and "first_public<=" in part.query for part in partitions)
    assert all(' marine="' not in f" {part.query}" for part in partitions)


def test_ena_portal_run_parsing_and_accessibility_statuses(tmp_path, monkeypatch):
    assert classify_sequence_accessibility({"fastq_ftp": "ftp.sra/x.fastq.gz", "fastq_bytes": "123"}) == "fastq_confirmed"
    assert classify_sequence_accessibility({"submitted_ftp": "ftp.sra/x.bam", "submitted_format": "BAM"}) == "submitted_reads_confirmed"
    assert classify_sequence_accessibility({"sra_ftp": "ftp.sra/x.sra", "sra_bytes": "10"}) == "sra_archive_confirmed"
    assert classify_sequence_accessibility({"fastq_ftp": "ftp.sra/x.fastq.gz"}) == "sequence_locator_present_unverified"

    row = {
        "run_accession": "ERR1",
        "experiment_accession": "ERX1",
        "sample_accession": "SAMEA1",
        "study_accession": "PRJEB1",
        "secondary_study_accession": "ERP1",
        "study_title": "Marine sediment microbial communities",
        "fastq_ftp": "ftp.sra/x.fastq.gz",
        "fastq_bytes": "100",
        "library_strategy": "AMPLICON",
        "environment_biome": "marine sediment",
        "marine": "marine:high_confidence",
    }
    run = parse_read_run(row, marine_confidence="high", marine_match_methods="ena_marine_tag:marine:high_confidence")

    assert run is not None
    assert run.run_accession == "ERR1"
    assert run.bioproject_accession == "PRJEB1"
    assert run.sequence_accessibility_status == "fastq_confirmed"


def test_ena_portal_client_searches_read_runs(tmp_path, monkeypatch):
    db = _db(tmp_path)
    config = SeedDiscoveryConfig(db_path=tmp_path / "seeds.sqlite")
    http = CachedHttpClient(config, db)
    observed = {}

    def fake_get_json(source, url, params=None, use_cache=True):  # noqa: ANN001
        observed.update({"source": source, "url": url, "params": params})
        return [{"run_accession": "ERR1", "study_accession": "PRJEB1", "fastq_ftp": "ftp/x.fq.gz", "fastq_bytes": "1"}]

    monkeypatch.setattr(http, "get_json", fake_get_json)
    runs = EnaPortalClient(http, config).search_read_runs(
        'marine="marine:high_confidence"',
        marine_confidence="high",
        marine_match_methods="tag",
    )

    assert observed["source"] == "ena_portal"
    assert observed["params"]["result"] == "read_run"
    assert "offset" not in observed["params"]
    assert runs[0].run_accession == "ERR1"


def test_ena_study_aggregation_and_paper_seed_view(tmp_path):
    db = _db(tmp_path)
    db.upsert_ena_run(
        EnaRun(
            run_accession="ERR1",
            study_accession="PRJEB1",
            secondary_study_accession="ERP1",
            bioproject_accession="PRJEB1",
            sample_accession="SAMEA1",
            study_title="Marine sediment microbial communities",
            fastq_ftp="ftp/x_1.fastq.gz",
            fastq_bytes="100",
            collection_date="2020-01-01",
            lat="1.0",
            lon="2.0",
            depth="5",
            environment_biome="marine sediment",
            target_gene="16S",
            library_strategy="AMPLICON",
            sequence_accessibility_status="fastq_confirmed",
            marine_confidence="high",
            marine_match_methods="ena_marine_tag:marine:high_confidence",
            marine_tag="marine:high_confidence",
        )
    )
    db.upsert_ena_run(
        EnaRun(
            run_accession="ERR2",
            study_accession="PRJEB1",
            secondary_study_accession="ERP1",
            bioproject_accession="PRJEB1",
            sample_accession="SAMEA2",
            study_title="Marine sediment microbial communities",
            submitted_ftp="ftp/x_2.bam",
            submitted_format="BAM",
            sequence_accessibility_status="submitted_reads_confirmed",
            marine_confidence="medium",
            marine_match_methods="ena_marine_tag:marine:medium_confidence",
        )
    )
    groups = db.ena_run_groups()
    study = aggregate_ena_study(next(iter(groups.values())))
    ena_study_id = db.upsert_ena_study(study)

    assert study.canonical_dataset_id == "PRJEB1"
    assert study.sample_count == 2
    assert study.downloadable_run_count == 2
    assert study.sequence_accessibility_status == "fastq_confirmed"
    assert study.metadata_usefulness_score >= 5

    db.upsert_ena_publication_candidate(
        ena_study_id,
        PublicationCandidate(
            doi="10.1000/ena",
            title="Marine sediment microbial communities",
            publication_date="2021-01-01",
            publication_year=2021,
            match_method="ena_xref",
            match_confidence=MatchConfidence.VERY_HIGH,
        ),
    )
    primary_id, status, reason = choose_primary_candidate(db.candidates_for_ena_study(ena_study_id))
    db.set_ena_primary(ena_study_id, primary_id, status, reason)
    row = db.conn.execute("SELECT * FROM paper_seeds WHERE seed_source = 'ena' AND canonical_dataset_id = 'PRJEB1'").fetchone()

    assert row["primary_doi"] == "10.1000/ena"
    assert row["seed_status"] == "complete"


def test_publication_candidate_deduplication_and_primary_selection(tmp_path):
    db = _db(tmp_path)
    study_id = db.upsert_study(MgnifyStudy(mgnify_accession="MGYS0001", biome="Marine"))

    db.upsert_publication_candidate(
        study_id,
        PublicationCandidate(
            doi="https://doi.org/10.1000/LATER",
            title="Later paper",
            publication_date="2022-01-01",
            publication_year=2022,
            match_method="mgnify_publication",
            match_confidence=MatchConfidence.VERY_HIGH,
        ),
    )
    db.upsert_publication_candidate(
        study_id,
        PublicationCandidate(
            doi="doi:10.1000/EARLY",
            title="Early paper",
            publication_date="2020-01-01",
            publication_year=2020,
            match_method="ena_xref",
            match_confidence=MatchConfidence.VERY_HIGH,
        ),
    )
    db.upsert_publication_candidate(
        study_id,
        PublicationCandidate(
            doi="10.1000/early",
            title="Early paper duplicate",
            publication_date="2020-01-01",
            publication_year=2020,
            match_method="openalex_accession",
            match_confidence=MatchConfidence.HIGH,
        ),
    )

    candidates = db.candidates_for_study(study_id)
    assert len(candidates) == 2
    primary_id, status, reason = choose_primary_candidate(candidates)
    db.set_primary(study_id, primary_id, status, reason)

    row = db.conn.execute("SELECT * FROM paper_seeds WHERE mgnify_accession = 'MGYS0001'").fetchone()
    assert row["primary_doi"] == "10.1000/early"
    assert row["publication_candidate_count"] == 2


def test_primary_selection_does_not_resolve_title_only_candidate(tmp_path):
    db = _db(tmp_path)
    study_id = db.upsert_study(MgnifyStudy(mgnify_accession="MGYS_TITLE_ONLY", biome="Marine"))
    db.upsert_publication_candidate(
        study_id,
        PublicationCandidate(
            title="A title-only MGnify publication relationship",
            match_method="mgnify_publication",
            match_confidence=MatchConfidence.VERY_HIGH,
        ),
    )

    primary_id, status, reason = choose_primary_candidate(db.candidates_for_study(study_id))

    assert primary_id is None
    assert status.value == "publication_candidates_low_confidence"
    assert "none had DOI" in (reason or "")


def test_resolution_resume_skips_no_publication_until_refresh(tmp_path):
    db = _db(tmp_path)
    no_pub_id = db.upsert_study(MgnifyStudy(mgnify_accession="MGYS_NO_PUB", biome="Marine"))
    pending_id = db.upsert_study(MgnifyStudy(mgnify_accession="MGYS_PENDING", biome="Marine"))
    db.set_primary(no_pub_id, None, status=ResolutionStatus.NO_PUBLICATION)

    rows = db.studies_for_resolution(refresh=False, limit=10)
    refresh_rows = db.studies_for_resolution(refresh=True, limit=10)

    assert [row["id"] for row in rows] == [pending_id]
    assert {row["id"] for row in refresh_rows} >= {no_pub_id, pending_id}


def test_resolution_queue_prioritizes_least_recently_checked_reprocess_rows(tmp_path):
    db = _db(tmp_path)
    old_id = db.upsert_study(MgnifyStudy(mgnify_accession="MGYS_OLD", biome="Marine"))
    new_id = db.upsert_study(MgnifyStudy(mgnify_accession="MGYS_NEW", biome="Marine"))
    db.conn.execute(
        "UPDATE mgnify_studies SET publication_resolution_status = 'openalex_no_resolve_reprocess', last_checked_at = '2026-08-22T10:00:00+00:00' WHERE id = ?",
        (old_id,),
    )
    db.conn.execute(
        "UPDATE mgnify_studies SET publication_resolution_status = 'openalex_no_resolve_reprocess', last_checked_at = '2026-08-22T12:00:00+00:00' WHERE id = ?",
        (new_id,),
    )
    db.conn.commit()

    rows = db.studies_for_resolution(refresh=False, limit=2)

    assert [row["mgnify_accession"] for row in rows] == ["MGYS_OLD", "MGYS_NEW"]


def test_mgnify_pagination_uses_count_items_shape(tmp_path, monkeypatch):
    db = _db(tmp_path)
    config = SeedDiscoveryConfig(db_path=tmp_path / "seeds.sqlite", page_size=2)
    http = CachedHttpClient(config, db)

    calls: list[str] = []

    def fake_get(url, params=None):  # noqa: ANN001
        calls.append(f"{url}?page={params['page']}")
        return httpx.Response(
            200,
            request=httpx.Request("GET", url),
            json={
                "count": 3,
                "items": [{"accession": f"MGYS{params['page']}A"}, {"accession": f"MGYS{params['page']}B"}]
                if params["page"] == 1
                else [{"accession": "MGYS2A"}],
            },
        )

    monkeypatch.setattr(http._client, "get", fake_get)
    client = MgnifyClient(http, config)

    pages = list(client.iter_study_payloads(max_pages=2))

    assert [page for page, _payload in pages] == [1, 1, 2]
    assert calls == [f"{config.mgnify_base_url}/studies/?page=1", f"{config.mgnify_base_url}/studies/?page=2"]


def test_ena_xref_parsing_returns_europepmc_candidates(tmp_path, monkeypatch):
    db = _db(tmp_path)
    config = SeedDiscoveryConfig(db_path=tmp_path / "seeds.sqlite")
    http = CachedHttpClient(config, db)

    monkeypatch.setattr(
        http,
        "get_json",
        lambda *args, **kwargs: [
            {
                "Source": "EuropePMC",
                "Source primary accession": "PMC123",
                "Source secondary accession": "98765",
            }
        ],
    )

    candidates = EnaXrefClient(http, config).publications_for_accession("PRJEB1")

    assert candidates[0].pmcid == "PMC123"
    assert candidates[0].pmid == "98765"
    assert candidates[0].match_method == "ena_xref"


def test_europepmc_and_openalex_accession_search_parsing(tmp_path, monkeypatch):
    db = _db(tmp_path)
    config = SeedDiscoveryConfig(db_path=tmp_path / "seeds.sqlite")
    http = CachedHttpClient(config, db)

    def fake_get_json(source, url, params=None, use_cache=True):  # noqa: ANN001
        if source == "europepmc":
            return {
                "resultList": {
                    "result": [{"doi": "10.1234/test", "pmid": "123", "title": "Paper", "pubYear": "2021"}]
                }
            }
        return {
            "results": [
                {
                    "id": "https://openalex.org/W123",
                    "doi": "https://doi.org/10.5678/open",
                    "title": "PRJEB1 study paper",
                    "publication_year": 2020,
                }
            ]
        }

    monkeypatch.setattr(http, "get_json", fake_get_json)

    europe = EuropePmcSeedClient(http, config).accession_search("PRJEB1")
    openalex = OpenAlexSeedClient(http, config).accession_search("PRJEB1")

    assert europe[0].doi == "10.1234/test"
    assert openalex[0].openalex_id == "W123"
    assert openalex[0].match_method == "openalex_accession"


def test_openalex_requests_include_contact_email_and_optional_api_key(tmp_path, monkeypatch):
    db = _db(tmp_path)
    # openalex_mailto pinned explicitly rather than relying on whichever
    # email SeedDiscoveryConfig's own FAIR_OCEAN_CONTACT_EMAIL-driven
    # default happens to resolve to in this environment (see config.py) --
    # this test is about the mailto param being sent at all, not about
    # any one specific address.
    config = SeedDiscoveryConfig(
        db_path=tmp_path / "seeds.sqlite", openalex_api_key="KEY", openalex_mailto="pinned@example.com"
    )
    http = CachedHttpClient(config, db)
    observed_params = {}

    def fake_get_json(source, url, params=None, use_cache=True):  # noqa: ANN001
        observed_params.update(params or {})
        return {"results": []}

    monkeypatch.setattr(http, "get_json", fake_get_json)

    OpenAlexSeedClient(http, config).accession_search("PRJEB1")

    assert observed_params["mailto"] == "pinned@example.com"
    assert observed_params["api_key"] == "KEY"


def test_http_client_raises_openalex_429_without_retry_after_sleep(tmp_path, monkeypatch):
    db = _db(tmp_path)
    config = SeedDiscoveryConfig(db_path=tmp_path / "seeds.sqlite")
    http = CachedHttpClient(config, db)
    sleeps: list[float] = []

    def fake_sleep(seconds):  # noqa: ANN001
        sleeps.append(seconds)

    def fake_get(url, params=None):  # noqa: ANN001
        return httpx.Response(
            429,
            headers={"Retry-After": "21926"},
            request=httpx.Request("GET", url),
            json={"error": "too many requests"},
        )

    monkeypatch.setattr("fair_ocean_agent.seed_discovery.clients.http.time.sleep", fake_sleep)
    monkeypatch.setattr(http._client, "get", fake_get)

    try:
        try:
            http.get_json("openalex", "https://api.openalex.org/works", params={"search.exact": "PRJDB4107"}, use_cache=False)
        except httpx.HTTPStatusError as exc:
            assert exc.response.status_code == 429
        else:
            raise AssertionError("expected OpenAlex 429 to raise")
    finally:
        http.close()

    assert 21926 not in sleeps


def test_http_client_persists_openalex_interval_across_clients(tmp_path, monkeypatch):
    db = _db(tmp_path)
    config = SeedDiscoveryConfig(db_path=tmp_path / "seeds.sqlite")
    db.update_crawl_state("openalex", status="running")
    http = CachedHttpClient(config, db)
    sleeps: list[float] = []

    def fake_sleep(seconds):  # noqa: ANN001
        sleeps.append(seconds)

    def fake_get(url, params=None):  # noqa: ANN001
        return httpx.Response(
            200,
            request=httpx.Request("GET", url),
            json={"results": []},
        )

    monkeypatch.setattr("fair_ocean_agent.seed_discovery.clients.http.time.sleep", fake_sleep)
    monkeypatch.setattr(http._client, "get", fake_get)

    try:
        http.get_json("openalex", "https://api.openalex.org/works", params={"search.exact": "PRJDB4107"}, use_cache=False)
    finally:
        http.close()

    assert sleeps
    assert sleeps[0] > 9.0


def test_local_epmc_accession_index_resolves_direct_bioproject(tmp_path):
    db = _db(tmp_path)
    db.upsert_epmc_article_ids(
        [
            {
                "pmid": "30179232",
                "pmcid": "PMC6122167",
                "doi": "10.1000/project",
                "snapshot_date": "2026-08-22",
                "source_file": "PMID_PMCID_DOI.csv.gz",
            }
        ]
    )
    db.upsert_epmc_accession_links(
        [
            {
                "database_name": "bioproject",
                "accession": "PRJNA385854",
                "normalized_accession": "PRJNA385854",
                "pmcid": "PMC6122167",
                "article_source": "MED",
                "article_external_id": "30179232",
                "snapshot_date": "2026-08-22",
                "source_file": "bioproject.csv",
            }
        ]
    )

    candidates = LocalEuropePmcResolver(db).resolve_publication_for_dataset(DatasetAccessions(bioproject="prjna385854."))

    assert candidates[0].doi == "10.1000/project"
    assert candidates[0].match_confidence == MatchConfidence.VERY_HIGH
    assert candidates[0].match_method == "europe_pmc_bulk_accessions"


def test_local_epmc_accession_index_aggregates_biosample_evidence(tmp_path):
    db = _db(tmp_path)
    db.upsert_epmc_article_ids(
        [
            {
                "pmid": "30179232",
                "pmcid": "PMC6122167",
                "doi": "10.1000/samples",
                "snapshot_date": "2026-08-22",
                "source_file": "PMID_PMCID_DOI.csv.gz",
            }
        ]
    )
    db.upsert_epmc_accession_links(
        [
            {
                "database_name": "biosample",
                "accession": accession,
                "normalized_accession": accession,
                "pmcid": "PMC6122167",
                "article_source": "MED",
                "article_external_id": "30179232",
                "snapshot_date": "2026-08-22",
                "source_file": "biosample.csv",
            }
            for accession in ("SAMEA111", "SAMEA112", "SAMEA113", "SAMEA114")
        ]
    )

    candidates = LocalEuropePmcResolver(db).resolve_publication_for_dataset(
        DatasetAccessions(biosamples=["SAMEA111", "SAMEA112", "SAMEA113", "SAMEA114"])
    )

    assert candidates[0].doi == "10.1000/samples"
    assert candidates[0].match_confidence == MatchConfidence.VERY_HIGH
    assert candidates[0].match_score == 12.0


def test_publication_resolver_uses_ena_run_level_local_epmc_evidence(tmp_path):
    db = _db(tmp_path)
    db.upsert_ena_run(
        EnaRun(
            run_accession="ERR111",
            experiment_accession="ERX111",
            sample_accession="SAMEA111",
            study_accession="PRJEB_NO_HIT",
            secondary_study_accession="ERP_NO_HIT",
            bioproject_accession="PRJEB_NO_HIT",
            sequence_accessibility_status="fastq_confirmed",
        )
    )
    db.upsert_ena_run(
        EnaRun(
            run_accession="ERR112",
            experiment_accession="ERX112",
            sample_accession="SAMEA112",
            study_accession="PRJEB_NO_HIT",
            secondary_study_accession="ERP_NO_HIT",
            bioproject_accession="PRJEB_NO_HIT",
            sequence_accessibility_status="fastq_confirmed",
        )
    )
    ena_study_id = db.upsert_ena_study(
        aggregate_ena_study(
            [
                {
                    "run_accession": "ERR111",
                    "study_accession": "PRJEB_NO_HIT",
                    "secondary_study_accession": "ERP_NO_HIT",
                    "bioproject_accession": "PRJEB_NO_HIT",
                    "sample_accession": "SAMEA111",
                    "secondary_sample_accession": None,
                    "study_title": "Marine sediment amplicon sequencing",
                    "project_name": None,
                    "centre_name": None,
                    "first_public": None,
                    "sequence_accessibility_status": "fastq_confirmed",
                    "fastq_bytes": "1",
                    "marine_confidence": "high",
                    "marine_match_methods": "test",
                    "marine_tag": None,
                    "collection_date": None,
                    "lat": None,
                    "lon": None,
                    "depth": None,
                    "environment_biome": None,
                    "environment_feature": None,
                    "environment_material": None,
                    "sample_collection": None,
                    "target_gene": None,
                    "extraction_protocol": None,
                    "library_construction_protocol": None,
                    "library_strategy": "AMPLICON",
                    "library_source": None,
                }
            ]
        )
    )
    db.upsert_epmc_accession_links(
        [
            {
                "database_name": "gen",
                "accession": accession,
                "normalized_accession": accession,
                "pmcid": "PMC_RUNS",
                "article_source": "MED",
                "article_external_id": "12345",
                "snapshot_date": "2026-08-22",
                "source_file": "gen.csv",
            }
            for accession in ("ERR111", "ERR112")
        ]
    )

    row = db.conn.execute("SELECT * FROM ena_studies WHERE id = ?", (ena_study_id,)).fetchone()
    status = _resolver(db).resolve_ena_study(row)
    candidate = db.conn.execute(
        "SELECT * FROM publication_candidates WHERE ena_study_id = ? AND match_method = 'europe_pmc_bulk_accessions'",
        (ena_study_id,),
    ).fetchone()

    assert status == ResolutionStatus.RESOLVED
    assert candidate["pmid"] == "12345"
    assert candidate["match_confidence"] == MatchConfidence.HIGH.value


def test_publication_resolver_keeps_local_epmc_candidate_when_pmid_enrichment_fails(tmp_path):
    db = _db(tmp_path)
    ena_study_id = db.upsert_ena_study(
        aggregate_ena_study(
            [
                {
                    "run_accession": "ERR111",
                    "study_accession": "PRJEB_NO_HIT",
                    "secondary_study_accession": "ERP_NO_HIT",
                    "bioproject_accession": "PRJEB_NO_HIT",
                    "sample_accession": "SAMEA111",
                    "secondary_sample_accession": None,
                    "study_title": "Marine sediment amplicon sequencing",
                    "project_name": None,
                    "centre_name": None,
                    "first_public": None,
                    "sequence_accessibility_status": "fastq_confirmed",
                    "fastq_bytes": "1",
                    "marine_confidence": "high",
                    "marine_match_methods": "test",
                    "marine_tag": None,
                    "collection_date": None,
                    "lat": None,
                    "lon": None,
                    "depth": None,
                    "environment_biome": None,
                    "environment_feature": None,
                    "environment_material": None,
                    "sample_collection": None,
                    "target_gene": None,
                    "extraction_protocol": None,
                    "library_construction_protocol": None,
                    "library_strategy": "AMPLICON",
                    "library_source": None,
                }
            ]
        )
    )
    db.upsert_ena_run(
        EnaRun(
            run_accession="ERR111",
            study_accession="PRJEB_NO_HIT",
            secondary_study_accession="ERP_NO_HIT",
            bioproject_accession="PRJEB_NO_HIT",
            sequence_accessibility_status="fastq_confirmed",
        )
    )
    db.upsert_epmc_accession_links(
        [
            {
                "database_name": "gen",
                "accession": "ERR111",
                "normalized_accession": "ERR111",
                "pmcid": None,
                "article_source": "MED",
                "article_external_id": "12345",
                "snapshot_date": "2026-08-22",
                "source_file": "gen.csv",
            }
        ]
    )

    class BrokenEuropePmc(_EmptyEuropePmc):
        def resolve_pmid(self, pmid):  # noqa: ANN001
            raise RuntimeError("network down")

    row = db.conn.execute("SELECT * FROM ena_studies WHERE id = ?", (ena_study_id,)).fetchone()
    status = _resolver(db, europepmc=BrokenEuropePmc()).resolve_ena_study(row)
    candidate = db.conn.execute("SELECT * FROM publication_candidates WHERE ena_study_id = ?", (ena_study_id,)).fetchone()

    assert status == ResolutionStatus.RESOLVED
    assert candidate["pmid"] == "12345"
    assert candidate["normalized_doi"] is None


def test_epmc_accession_normalization_trims_punctuation():
    assert normalize_epmc_accession("(prjeb12345).") == "PRJEB12345"


def test_epmc_id_mapping_rejects_error_payload(tmp_path):
    path = tmp_path / "PMID_PMCID_DOI.csv.gz"
    with gzip.open(path, "wt", encoding="utf-8") as handle:
        handle.write("ERROR:\nORA-12537: TNS:connection closed\n")

    try:
        list(parse_id_mapping(path, snapshot_date="2026-08-22"))
    except RuntimeError as exc:
        assert "error payload" in str(exc)
    else:
        raise AssertionError("expected RuntimeError")


class _EmptyMgnify:
    def publications(self, accession):  # noqa: ANN001
        return []


class _EmptyEna:
    def publications_for_accession(self, accession):  # noqa: ANN001
        return []


class _EmptyNcbi:
    def pubmed_for_bioproject(self, accession):  # noqa: ANN001
        return []


class _EmptyOpenAlex:
    def accession_search(self, accession):  # noqa: ANN001
        return []

    def title_search(self, title):  # noqa: ANN001
        return []

    def metadata_search(self, query):  # noqa: ANN001
        return []


class _EmptyEuropePmc:
    def accession_search(self, accession):  # noqa: ANN001
        return []

    def title_search(self, title):  # noqa: ANN001
        return []

    def resolve_pmid(self, pmid):  # noqa: ANN001
        return None


def _resolver(db, config=None, **clients):  # noqa: ANN001
    return PublicationResolver(
        db,
        config or SeedDiscoveryConfig(),
        mgnify=clients.get("mgnify") or _EmptyMgnify(),
        ena=clients.get("ena") or _EmptyEna(),
        ncbi=clients.get("ncbi") or _EmptyNcbi(),
        openalex=clients.get("openalex") or _EmptyOpenAlex(),
        europepmc=clients.get("europepmc") or _EmptyEuropePmc(),
    )


def test_title_similarity_accepts_publication_like_title_variants():
    score = title_similarity(
        "Metabolically active microbial communities in marine sediment under high-CO2 and low-pH extremes",
        "Metabolically active microbial communities in marine sediment under high CO2 and low pH extremes",
    )

    assert score >= 0.98


def test_publication_resolver_uses_europepmc_title_search_when_accessions_fail(tmp_path):
    db = _db(tmp_path)
    study_id = db.upsert_study(
        MgnifyStudy(
            mgnify_accession="MGYS_TITLE",
            bioproject_accession="PRJDB1",
            study_name="Metabolically active microbial communities in marine sediment under high-CO2 and low-pH extremes",
            biome="Marine sediment",
        )
    )

    class EuropeTitle(_EmptyEuropePmc):
        def title_search(self, title):  # noqa: ANN001
            return [
                PublicationCandidate(
                    doi="10.1000/title",
                    title="Metabolically active microbial communities in marine sediment under high CO2 and low pH extremes",
                    publication_date="2014-01-01",
                    publication_year=2014,
                    match_method="europepmc_title_search",
                    matched_identifier=title,
                    match_confidence=MatchConfidence.MEDIUM,
                )
            ]

    row = db.conn.execute("SELECT * FROM mgnify_studies WHERE id = ?", (study_id,)).fetchone()
    status = _resolver(db, europepmc=EuropeTitle()).resolve_study(row)

    seed = db.conn.execute("SELECT * FROM paper_seeds WHERE mgnify_accession = 'MGYS_TITLE'").fetchone()
    assert status == ResolutionStatus.RESOLVED
    assert seed["primary_doi"] == "10.1000/title"
    assert seed["publication_match_method"] == "europepmc_title_search"


def test_publication_resolver_raises_openalex_429_to_stop_run(tmp_path):
    db = _db(tmp_path)
    study_id = db.upsert_study(
        MgnifyStudy(
            mgnify_accession="MGYS_BLOCKED",
            bioproject_accession="PRJDB2",
            study_name="Diversity of bacterioneuston and bacterioplankton in coastal waters of Misaki Japan",
            biome="Coastal water",
        )
    )

    class BlockedOpenAlex(_EmptyOpenAlex):
        def accession_search(self, accession):  # noqa: ANN001
            response = httpx.Response(429, request=httpx.Request("GET", "https://api.openalex.org/works"))
            raise httpx.HTTPStatusError("too many requests", request=response.request, response=response)

        def title_search(self, title):  # noqa: ANN001
            response = httpx.Response(429, request=httpx.Request("GET", "https://api.openalex.org/works"))
            raise httpx.HTTPStatusError("too many requests", request=response.request, response=response)

    row = db.conn.execute("SELECT * FROM mgnify_studies WHERE id = ?", (study_id,)).fetchone()
    try:
        _resolver(db, config=SeedDiscoveryConfig(openalex_enabled=True), openalex=BlockedOpenAlex()).resolve_study(row)
    except OpenAlexRateLimitError as exc:
        assert "OpenAlex returned 429" in str(exc)
    else:
        raise AssertionError("expected OpenAlexRateLimitError")

    seed = db.conn.execute("SELECT * FROM paper_seeds WHERE mgnify_accession = 'MGYS_BLOCKED'").fetchone()
    assert seed["publication_resolution_status"] == "not_yet_processed"


def test_publication_resolver_can_disable_openalex_and_flag_publication_like_title(tmp_path):
    db = _db(tmp_path)
    study_id = db.upsert_study(
        MgnifyStudy(
            mgnify_accession="MGYS_NO_OPENALEX",
            bioproject_accession="PRJDB3",
            study_name="Diversity of microbial communities in marine sediment using amplicon sequencing",
            biome="Marine sediment",
        )
    )

    class FailingOpenAlex(_EmptyOpenAlex):
        def accession_search(self, accession):  # noqa: ANN001
            raise AssertionError("OpenAlex accession search should be disabled")

        def title_search(self, title):  # noqa: ANN001
            raise AssertionError("OpenAlex title search should be disabled")

        def metadata_search(self, query):  # noqa: ANN001
            raise AssertionError("OpenAlex metadata search should be disabled")

    row = db.conn.execute("SELECT * FROM mgnify_studies WHERE id = ?", (study_id,)).fetchone()
    status = _resolver(
        db,
        config=SeedDiscoveryConfig(openalex_enabled=False),
        openalex=FailingOpenAlex(),
    ).resolve_study(row)

    assert status == ResolutionStatus.OPENALEX_REPROCESS
    candidate = db.conn.execute(
        "SELECT * FROM publication_candidates WHERE mgnify_study_id = ?",
        (study_id,),
    ).fetchone()
    assert candidate["title"] == "Diversity of microbial communities in marine sediment using amplicon sequencing"
    assert candidate["match_method"] == "mgnify_publication_like_title"
    assert candidate["match_confidence"] == MatchConfidence.LOW.value


def test_publication_resolver_can_disable_openalex_for_ena_study(tmp_path):
    db = _db(tmp_path)
    ena_study_id = db.upsert_ena_study(
        aggregate_ena_study(
            [
                {
                    "run_accession": "ERR1",
                    "study_accession": "ERP1",
                    "secondary_study_accession": "ERP1",
                    "bioproject_accession": None,
                    "sample_accession": "ERS1",
                    "secondary_sample_accession": None,
                    "study_title": "Diversity of microbial communities in coastal sediment using amplicon sequencing",
                    "project_name": None,
                    "centre_name": "Example Centre",
                    "first_public": "2021-01-01",
                    "sequence_accessibility_status": "fastq_confirmed",
                    "fastq_bytes": "1",
                    "marine_confidence": "high",
                    "marine_match_methods": "ena_marine_tag:marine:high_confidence",
                    "marine_tag": "marine:high_confidence",
                    "collection_date": None,
                    "lat": None,
                    "lon": None,
                    "depth": None,
                    "environment_biome": None,
                    "environment_feature": None,
                    "environment_material": None,
                    "sample_collection": None,
                    "target_gene": None,
                    "extraction_protocol": None,
                    "library_construction_protocol": None,
                    "library_strategy": "AMPLICON",
                    "library_source": None,
                }
            ]
        )
    )

    class FailingOpenAlex(_EmptyOpenAlex):
        def accession_search(self, accession):  # noqa: ANN001
            raise AssertionError("OpenAlex accession search should be disabled")

        def title_search(self, title):  # noqa: ANN001
            raise AssertionError("OpenAlex title search should be disabled")

    row = db.conn.execute("SELECT * FROM ena_studies WHERE id = ?", (ena_study_id,)).fetchone()
    status = _resolver(
        db,
        config=SeedDiscoveryConfig(openalex_enabled=False),
        openalex=FailingOpenAlex(),
    ).resolve_ena_study(row)

    assert status == ResolutionStatus.OPENALEX_REPROCESS
    candidate = db.conn.execute(
        "SELECT * FROM publication_candidates WHERE ena_study_id = ?",
        (ena_study_id,),
    ).fetchone()
    assert candidate["match_method"] == "ena_publication_like_title"


def test_jgi_gold_snapshot_inspection_and_ingest(tmp_path):
    raw_dir = tmp_path / "jgi_gold" / "raw" / "2026-08-22"
    raw_dir.mkdir(parents=True)
    (raw_dir / "manifest.json").write_text('{"snapshot": "2026-08-22", "files": []}', encoding="utf-8")

    workbook = Workbook()
    studies = workbook.active
    studies.title = "Studies"
    studies.append(["GOLD Study ID", "Study Name", "Study Description", "NCBI BioProject Accession", "Publication"])
    studies.append(["Gs0001", "Marine sediment metagenome", "Coastal ocean sediment survey", "PRJNA000001", "doi:10.1234/example"])

    biosamples = workbook.create_sheet("Biosamples")
    biosamples.append(
        [
            "GOLD Biosample ID",
            "GOLD Study ID",
            "NCBI Biosample Accession",
            "Biosample Name",
            "Sample Collection Date",
            "Latitude",
            "Longitude",
            "Depth",
            "Ecosystem",
            "Ecosystem Category",
            "Ecosystem Type",
            "Ecosystem Subtype",
            "Specific Ecosystem",
            "Environmental Medium",
            "Sample Collection Method",
            "Size Fraction",
        ]
    )
    biosamples.append(
        [
            "Gb0001",
            "Gs0001",
            "SAMN000001",
            "Marine mud 1",
            "2020-01-02",
            "42.1",
            "-70.2",
            "15 m",
            "Environmental",
            "Aquatic",
            "Marine",
            "Marine sediment",
            "Coastal sediment",
            "sediment",
            "box corer",
            "0.2-3 um",
        ]
    )

    projects = workbook.create_sheet("Sequencing Projects")
    projects.append(
        [
            "GOLD Sequencing Project ID",
            "GOLD Study ID",
            "GOLD Biosample ID",
            "NCBI BioProject Accession",
            "NCBI Biosample Accession",
            "Sequencing Strategy",
            "Project Status",
            "JGI Project ID",
        ]
    )
    projects.append(["Gp0001", "Gs0001", "Gb0001", "PRJNA000001", "SAMN000001", "Metagenome", "Public", "JGI123"])

    analyses = workbook.create_sheet("Analysis Projects")
    analyses.append(["GOLD Analysis Project ID", "GOLD Project ID", "GOLD Biosample ID", "GOLD Study ID", "Analysis Project Type", "IMG Taxon ID"])
    analyses.append(["Ga0001", "Gp0001", "Gb0001", "Gs0001", "IMG/M metagenome analysis", "3300000001"])

    workbook.save(raw_dir / "public_studies_biosamples_sps_aps_organisms.xlsx")
    _force_xlsx_a1_dimensions(raw_dir / "public_studies_biosamples_sps_aps_organisms.xlsx")

    inventory = inspect_gold_snapshot(raw_dir, tmp_path / "jgi_gold" / "processed" / "inspect-only")
    assert inventory["workbooks"]["public_studies_biosamples_sps_aps_organisms.xlsx"]["sheets"]["Biosamples"]["inferred_entity_type"] == "biosample"

    config = SeedDiscoveryConfig(db_path=tmp_path / "seeds.sqlite", gold_data_dir=tmp_path / "jgi_gold")
    locations = process_gold_snapshot(config, raw_dir, snapshot="2026-08-22")

    db = SeedDiscoveryDB(config.db_path)
    try:
        assert db.conn.execute("SELECT count(*) AS n FROM gold_source_rows").fetchone()["n"] == 4
        assert db.conn.execute("SELECT primary_bioproject FROM gold_studies WHERE gold_study_id = 'Gs0001'").fetchone()["primary_bioproject"] == "PRJNA000001"
        sample = db.conn.execute("SELECT * FROM gold_biosamples WHERE gold_biosample_id = 'Gb0001'").fetchone()
        assert sample["ncbi_biosample_accession"] == "SAMN000001"
        assert sample["marine_confidence"] == "high"
        staged = db.conn.execute("SELECT * FROM gold_faire_enrichment WHERE canonical_biosample = 'SAMN000001'").fetchone()
        assert staged["decimalLatitude"] == "42.1"
        assert staged["env_medium"] == "sediment"
        assert db.conn.execute("SELECT jgi_project_id FROM gold_sequencing_projects WHERE gold_project_id = 'Gp0001'").fetchone()["jgi_project_id"] == "JGI123"
        assert db.conn.execute("SELECT img_identifier FROM gold_analysis_projects WHERE gold_analysis_project_id = 'Ga0001'").fetchone()["img_identifier"] == "3300000001"
        assert db.conn.execute("SELECT doi FROM gold_study_publications WHERE gold_study_id = 'Gs0001'").fetchone()["doi"] == "10.1234/example"
        assert db.conn.execute("SELECT availability_status FROM gold_project_jgi_files WHERE gold_project_id = 'Gp0001'").fetchone()["availability_status"] == "metadata_only_auth_required_for_file_listing"
    finally:
        db.close()

    assert Path(locations["schema_inventory"]).exists()
    assert Path(locations["reports"]["faire_mapping_candidates"]).exists()
    assert Path(locations["jgi_file_manifest"]).exists()
    assert Path(locations["reports"]["metadata_completeness"]).exists()
    assert (tmp_path / "jgi_gold" / "OUTPUT_LOCATIONS.json").exists()


def _force_xlsx_a1_dimensions(path: Path) -> None:
    rewritten = path.with_suffix(".rewritten.xlsx")
    with ZipFile(path, "r") as src, ZipFile(rewritten, "w", ZIP_DEFLATED) as dst:
        for name in src.namelist():
            payload = src.read(name)
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                payload = re.sub(rb'<dimension ref="[^"]+"', b'<dimension ref="A1"', payload, count=1)
            dst.writestr(name, payload)
    rewritten.replace(path)
